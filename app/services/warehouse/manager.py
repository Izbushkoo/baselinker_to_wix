from sqlmodel import SQLModel, Field, Session, create_engine, select
from sqlalchemy import Column, JSON, text, func
from sqlalchemy.ext.asyncio import create_async_engine
from sqlmodel.ext.asyncio.session import AsyncSession
from typing import List, Optional, Tuple, Dict, Any, Union
import pandas as pd
from io import BytesIO
from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Query
from fastapi.responses import JSONResponse, StreamingResponse
from datetime import datetime, date, timedelta
from enum import Enum
import logging
import openpyxl
from PIL import Image
import io

from app.celery_shared import celery
from app.models.warehouse import Product, Stock, Sale, Transfer
from app.core.config import settings
from app.services.operations_service import OperationsService, get_operations_service

logger = logging.getLogger(__name__)

class StockError(Exception):
    """Исключение для ошибок, связанных с остатками на складе."""
    pass

class Warehouses(str, Enum):
    """Перечисление доступных складов в системе.
    
    Attributes:
        A (str): Основной склад для списаний
        B (str): Основной склад источник, первым отображается в файле остатков
    """
    A = 'Ирина'  # Основной склад для списаний
    B = 'Женя'  # Основной склад источник, первым отображается в файле остатков

def get_manager() -> 'InventoryManager':
    dsn = settings.SQLALCHEMY_DATABASE_URI.unicode_string()
    async_dsn = settings.SQLALCHEMY_DATABASE_URI_ASYNC.unicode_string()
    return InventoryManager(dsn, async_dsn)

class InventoryManager:
    def __init__(self, dsn: str, async_dsn: str):
        '''Менеджер инвентаря через PostgreSQL / SQLModel.'''
        self.engine = create_engine(dsn)
        self.async_engine = create_async_engine(async_dsn)
        self.operations_service: OperationsService = get_operations_service()

    def restock(self, sku: str, warehouse: str, quantity: int):
        '''Пополнение остатков на указанном складе.'''
        if quantity <= 0:
            raise ValueError('Количество пополнения должно быть положительным.')
            # Используем значение склада как есть, без преобразования в ключ
        with Session(self.engine) as session:
            key = (sku, warehouse)
            stock = session.get(Stock, key)
            if not stock:
                stock = Stock(sku=sku, warehouse=warehouse, quantity=quantity)
            else:
                stock.quantity += quantity
            session.add(stock)
            session.commit()

    async def count_products(self) -> int:
        '''Подсчитывает общее количество уникальных товаров в базе данных.'''
        async with AsyncSession(self.async_engine) as session:
            query = select(func.count()).select_from(Product)
            result = await session.exec(query)
            return result.first()

    async def get_low_stock_products(self) -> int:
        '''Подсчитывает количество товаров с суммарным остатком от 1 до 5 штук.'''
        async with AsyncSession(self.async_engine) as session:
            query = (
                select(func.count())
                .select_from(
                    select(Product.sku)
                    .outerjoin(Stock, Stock.sku == Product.sku)
                    .group_by(Product.sku)
                    .having(
                        func.coalesce(func.sum(Stock.quantity), 0) > 0,
                        func.coalesce(func.sum(Stock.quantity), 0) < 6
                    )
                    .subquery()
                )
            )
            result = await session.exec(query)
            return result.first()

    def transfer(
        self,
        sku: str,
        source: str,
        destination: str,
        quantity: int,
    ) -> None:
        """
        Перемещает указанное количество товара с одного склада на другой.
        
        Args:
            sku: SKU товара
            source: Склад-источник
            destination: Склад-получатель
            quantity: Количество для перемещения
        
        Raises:
            ValueError: Если количество отрицательное или склады совпадают
            StockError: Если на складе-источнике недостаточно товара
        """
        logger.info(f"Начало перемещения {quantity} единиц товара {sku} со склада {source} на склад {destination}")
        
        with Session(self.engine) as session:
            try:
                if quantity <= 0:
                    raise ValueError("Количество должно быть положительным")
                
                if source == destination:
                    raise ValueError("Склады источника и назначения должны различаться")

                # Проверяем наличие товара на складе-источнике
                source_stock = session.get(Stock, (sku, source))
                if not source_stock:
                    raise StockError(f"Товар {sku} отсутствует на складе {source}")
                
                if source_stock.quantity < quantity:
                    raise StockError(
                        f"Недостаточно товара '{sku}' на складе '{source}'. "
                        f"Запрошено: {quantity}, доступно: {source_stock.quantity}"
                    )

                # Получаем или создаем запись о товаре на складе-получателе
                dest_stock = session.get(Stock, (sku, destination))
                if not dest_stock:
                    dest_stock = Stock(sku=sku, warehouse=destination, quantity=0)
                    session.add(dest_stock)

                # Обновляем количества
                source_stock.quantity -= quantity
                dest_stock.quantity += quantity
                
                # Создаем запись о перемещении
                transfer_log = Transfer(
                    sku=sku,
                    source=source,
                    destination=destination,
                    quantity=quantity
                )
                session.add(transfer_log)
                
                session.commit()
                logger.info(f"Успешно перемещено {quantity} единиц товара {sku} со склада {source} на склад {destination}")
                
            except Exception as e:
                session.rollback()
                logger.error(f"Ошибка при перемещении товара: {str(e)}")
                raise

    def _remove_from_warehouse_base(self, session: Session, sku: str, warehouse: str, quantity: int):
        '''Базовая логика списания количества со склада.'''
        if quantity <= 0:
            raise ValueError('Количество списания должно быть положительным.')
            
        stock = session.get(Stock, (sku, warehouse))
        if not stock:
            raise ValueError(f'Товар {sku} не найден на складе {warehouse}')
            
        if stock.quantity < quantity:
            raise ValueError(f'Недостаточно товара на складе {warehouse}. Запрошено: {quantity}, доступно: {stock.quantity}')

        stock.quantity -= quantity
        session.add(stock)
        session.commit()
        # Запускаем задачу синхронизации Allegro по имени (без импорта)
        celery.send_task('app.services.allegro.sync_tasks.sync_allegro_stock_single_product', args=[sku])
        
        return stock

    def remove_from_warehouse(self, sku: str, warehouse: str, quantity: int):
        '''Списание количества с указанного склада без логирования продажи.'''
        logger.info(f'Начало списания {quantity} единиц товара {sku} со склада {warehouse}')
        
        try:
            with Session(self.engine) as session:
                self._remove_from_warehouse_base(session, sku, warehouse, quantity)
                
                logger.info(f'Успешно списано {quantity} единиц товара {sku} со склада {warehouse}')
                
        except Exception as e:
            logger.error(f'Ошибка при списании товара: {str(e)}')
            raise

    def remove_as_sale(self, sku: str, warehouse: str, quantity: int):
        '''Списание количества с указанного склада с логированием продажи.'''
        logger.info(f'Начало списания продажи {quantity} единиц товара {sku} со склада {warehouse}')
        
        try:
            with Session(self.engine) as session:
                self._remove_from_warehouse_base(session, sku, warehouse, quantity)
                
                # Логируем продажу
                sale = Sale(sku=sku, warehouse=warehouse, quantity=quantity)
                session.add(sale)
                session.commit()
                
                logger.info(f'Успешно списана продажа {quantity} единиц товара {sku} со склада {warehouse}')
                
        except Exception as e:
            logger.error(f'Ошибка при списании продажи: {str(e)}')
            raise

    def remove_one(self, sku: str, warehouse: str):
        '''Списание одной единицы с указанного склада.'''
        self.remove_from_warehouse(sku, warehouse, 1)

    def compress_image(self, image_data: bytes) -> Optional[bytes]:
        """
        Сжимает изображение до размера 100x100 пикселей в формат WebP.
        Поддерживает все форматы, которые поддерживает Pillow.
        
        Args:
            image_data: Байты исходного изображения
            
        Returns:
            bytes: Сжатые байты изображения в формате WebP или None в случае ошибки
        """
        if not image_data:
            return None
            
        try:
            # Открываем изображение из байтов
            img = Image.open(io.BytesIO(image_data))
            
            # Конвертируем в RGB если изображение в RGBA
            if img.mode in ('RGBA', 'LA'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                background.paste(img, mask=img.split()[-1])
                img = background
            
            # Определяем размеры для сохранения пропорций
            target_size = (100, 100)
            img.thumbnail(target_size, Image.Resampling.LANCZOS)
            
            # Сохраняем в буфер в формате WebP с максимальным сжатием
            output = io.BytesIO()
            img.save(output, format='WEBP', quality=30, method=6)
            
            return output.getvalue()
            
        except Exception as e:
            logging.error(f"Ошибка при сжатии изображения: {str(e)}")
            return None

    def find_header_row(self, file_bytes: bytes, required_cols: list, max_rows: int = 10) -> int:
        """Автоматически определяет строку с заголовками по наличию нужных колонок."""
        for header in range(max_rows):
            df = pd.read_excel(BytesIO(file_bytes), header=header, nrows=1)
            columns = [str(col).strip().lower() for col in df.columns]
            if all(col.lower() in columns for col in required_cols):
                return header
        raise ValueError(
            f"Не найдена строка с заголовками для колонок {required_cols}. В файле есть варианты: {columns}"
        )

    def import_incoming_from_excel(
        self,
        file_bytes: bytes,
        warehouse: str,
        sku_col: str = 'sku',
        qty_col: str = 'Кол-во',
        ean_col: str = 'EAN',
        name_col: str = 'Name',
        image_col: str = 'Foto',
        header: int = None
    ) -> List[Dict]:
        '''Импорт прихода из XLSX (bytes): SKU, количество, EAN, название и изображение.
        
        Если в файле отсутствуют колонки EAN, название или изображение, 
        то выполняется только пополнение остатков по SKU и количеству.
        '''
        logging.info("Начало импорта Excel файла")
        
        # Загружаем Excel файл
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        ws = wb.active
        header_row = header + 1 if header is not None else 1
        headers = [str(cell.value) for cell in ws[header_row]]
        logging.info(f"Найдены заголовки: {headers}")
        
        # Проверяем наличие обязательных колонок
        required_cols = [sku_col, qty_col]
        try:
            col_indices = {
                'sku': headers.index(sku_col),
                'qty': headers.index(qty_col),
            }
        except ValueError as e:
            raise ValueError(f"Не найдены обязательные колонки: {sku_col}, {qty_col}")
        
        # Проверяем наличие дополнительных колонок для полного режима
        full_mode = True
        try:
            col_indices['ean'] = headers.index(ean_col)
            col_indices['name'] = headers.index(name_col)
            logging.info("Обнаружен полный режим импорта - все колонки присутствуют")
        except ValueError:
            full_mode = False
            logging.info("Обнаружен режим пополнения - отсутствуют колонки EAN или название")
        
        # Создаем словарь изображений по номерам строк (только в полном режиме)
        images_by_row = {}
        
        if full_mode and image_col:
            try:
                col_indices['image'] = headers.index(image_col)
                for img in ws._images:
                    coords = img.anchor._from
                    row = coords.row + 1
                    if row > (header + 1 if header is not None else 1):
                        try:
                            if hasattr(img.ref, 'getvalue'):
                                image_data = img.ref.getvalue()
                            elif hasattr(img.ref, '_data'):
                                image_data = img.ref._data()
                            elif hasattr(img.ref, 'content'):
                                image_data = img.ref.content
                            else:
                                continue
                                
                            # Сжимаем изображение перед сохранением
                            compressed_image = self.compress_image(image_data)
                            if compressed_image:
                                images_by_row[row] = compressed_image
                        except Exception:
                            continue
            except ValueError:
                logging.warning("Колонка изображений не найдена, изображения не будут обработаны")

        # Список для хранения информации о обработанных товарах
        processed_products = []

        # Обрабатываем каждую строку данных
        start_row = header + 2 if header is not None else 2
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            try:
                sku = str(row[col_indices['sku']].value)
                
                # Пропускаем строку если SKU пустой
                if not sku or sku.lower() == 'none':
                    continue
                    
                qty = int(row[col_indices['qty']].value)
                
                if full_mode:
                    # Полный режим - создаем/обновляем продукт
                    ean = str(row[col_indices['ean']].value)
                    name = str(row[col_indices['name']].value)
                    image_data = images_by_row.get(row_idx)
                    
                    with Session(self.engine) as session:
                        product = session.get(Product, sku)
                        if not product:
                            product = Product(sku=sku, eans=[ean], name=name, image=image_data)
                            session.add(product)
                        else:
                            if ean not in product.eans:
                                product.eans.append(ean)
                            product.name = name
                            if image_data:
                                product.image = image_data
                            session.add(product)
                        session.commit()
                    
                    # Добавляем информацию о обработанном товаре
                    processed_products.append({
                        "sku": sku,
                        "quantity": qty,
                        "ean": ean,
                        "name": name
                    })
                else:
                    # Режим пополнения - только обновляем остатки
                    processed_products.append({
                        "sku": sku,
                        "quantity": qty
                    })
                
                # В любом случае добавляем или обновляем остатки на складе
                self.restock(sku, warehouse, qty)
                
            except Exception as e:
                logging.warning(f"Ошибка при обработке строки {row_idx}: {str(e)}")
                continue
        
        mode_str = "полного импорта" if full_mode else "пополнения остатков"
        logging.info(f"Импорт Excel файла завершен в режиме {mode_str}")
        return processed_products

    def import_transfer_from_excel(
        self,
        file_bytes: bytes,
        from_warehouse: str,
        to_warehouse: str,
        sku_col: str = 'sku',
        qty_col: str = 'кол-во',
        header: int = None
    ) -> Tuple[pd.DataFrame, List[Dict]]:
        '''Импорт перемещения из XLSX (bytes): SKU и количество.'''
        # Автоопределение строки с заголовками
        if header is None:
            required_cols = [sku_col, qty_col]
            header = self.find_header_row(file_bytes, required_cols)
        df = pd.read_excel(BytesIO(file_bytes), header=header)
        
        # Создаем колонку для ошибок
        df['error'] = None
        
        # Создаем список для хранения индексов успешно обработанных строк
        successful_rows = []
        
        # Список для хранения информации о успешно обработанных товарах
        processed_products = []
        
        # Обрабатываем каждую строку
        for idx, row in df.iterrows():
            try:
                # Пропускаем строку если SKU или количество - nan
                if pd.isna(row[sku_col]) or pd.isna(row[qty_col]):
                    df.at[idx, 'error'] = 'Отсутствует SKU или количество'
                    continue
                    
                sku = str(row[sku_col])
                # Пропускаем если SKU пустой
                if not sku or sku.lower() == 'none':
                    df.at[idx, 'error'] = 'Пустой SKU'
                    continue
                    
                qty = int(row[qty_col])
                self.transfer(sku, from_warehouse, to_warehouse, qty)
                successful_rows.append(idx)
                
                # Добавляем информацию о успешно обработанном товаре
                processed_products.append({
                    "sku": sku,
                    "quantity": qty
                })
                
            except KeyError as e:
                df.at[idx, 'error'] = f'Колонка не найдена: {e}'
            except ValueError as e:
                df.at[idx, 'error'] = str(e)
            except Exception as e:
                df.at[idx, 'error'] = f'Неизвестная ошибка: {str(e)}'
        
        # Удаляем успешно обработанные строки
        df = df.drop(successful_rows)
        
        # Если есть строки с ошибками, возвращаем их
        if not df.empty:
            return df, processed_products
        return pd.DataFrame(), processed_products

    def compress_all_product_images(self) -> Dict[str, bool]:
        """
        Сжимает все изображения продуктов в базе данных.
        
        Returns:
            Dict[str, bool]: Словарь с результатами сжатия для каждого SKU
                           {sku: True/False}, где True означает успешное сжатие
        """
        results = {}
        
        with Session(self.engine) as session:
            # Получаем все продукты с изображениями
            products = session.exec(select(Product).where(Product.image != None)).all()
            
            for product in products:
                try:
                    if not product.image:
                        results[product.sku] = False
                        continue
                        
                    # Сжимаем изображение
                    compressed_image = self.compress_image(product.image)
                    
                    if compressed_image:
                        # Обновляем изображение в базе
                        product.image = compressed_image
                        session.add(product)
                        results[product.sku] = True
                    else:
                        results[product.sku] = False
                        
                except Exception as e:
                    logging.error(f"Ошибка при сжатии изображения для SKU {product.sku}: {str(e)}")
                    results[product.sku] = False
            
            # Сохраняем все изменения
            session.commit()
            
        return results

    def get_stock_report(self) -> Tuple[pd.DataFrame, List[bytes]]:
        """
        Возвращает кортеж:
        - DataFrame с колонками sku, name, eans, warehouse_*, total (без image)
        - Список байтов изображений в том же порядке, что и строки в DataFrame
        """
        with Session(self.engine) as session:
            products = session.exec(select(Product)).all()
            stocks = session.exec(select(Stock)).all()

        if not products:
            empty_df = pd.DataFrame(columns=[
                'sku', 'name', 'eans', 'total'
            ] + [f'warehouse_{w.value}' for w in Warehouses])
            return empty_df, []

        # накапливаем строки
        rows = []
        images = []
        # создаём словарь stock_data с значениями складов
        stock_data = {s.sku: {w.value: 0 for w in Warehouses} for s in stocks}
        for s in stocks:
            stock_data.setdefault(s.sku, {w.value: 0 for w in Warehouses})
            stock_data[s.sku][s.warehouse] = s.quantity

        for prod in products:
            # Собираем общие данные (без image)
            row = {
                'sku': prod.sku,
                'name': prod.name,
                'eans': ', '.join(prod.eans) if prod.eans else '',
            }
            total = 0
            for wh in Warehouses:
                qty = stock_data.get(prod.sku, {}).get(wh.value, 0)
                row[f'warehouse_{wh.value}'] = qty
                total += qty
            row['total'] = total

            rows.append(row)
            # А картинку — собираем параллельно в отдельный список
            images.append(prod.image or b'')

        # Теперь df без колонки image
        df = pd.DataFrame(rows)
        return df, images

    def get_sales_report(self, start_date: date, end_date: date, sku: Optional[str] = None) -> pd.DataFrame:
        '''Получить DataFrame агрегированных логов продаж за период по SKU.'''
        with Session(self.engine) as session:
            stmt = select(Sale).where(Sale.timestamp >= datetime.combine(start_date, datetime.min.time()),
                                     Sale.timestamp <= datetime.combine(end_date, datetime.max.time()))
            if sku:
                # Разбиваем строку SKU на список и очищаем от пробелов
                sku_list = [s.strip() for s in sku.split(',') if s.strip()]
                if sku_list:
                    stmt = stmt.where(Sale.sku.in_(sku_list))
            sales = session.exec(stmt).all()
            
        rows = [{'sku': s.sku, 'warehouse': s.warehouse, 'quantity': s.quantity} for s in sales]
        df = pd.DataFrame(rows)
        
        # Агрегируем данные по SKU и складу
        if not df.empty:
            df = df.groupby(['sku', 'warehouse'])['quantity'].sum().reset_index()
            df = df.sort_values(['warehouse', 'sku'])
            
        return df

    def get_stock_by_sku(self, sku: str) -> dict:
        '''Получить остатки по конкретному SKU на всех складах.'''
        with Session(self.engine) as session:
            stocks = session.exec(select(Stock).where(Stock.sku == sku)).all()
            return {stock.warehouse: stock.quantity for stock in stocks}

    def clear_all_data(self):
        '''Удаляет все продукты, остатки и логи продаж из базы данных.'''
        with Session(self.engine) as session:
            session.exec(text("DELETE FROM transfer"))
            session.exec(text("DELETE FROM sale"))
            session.exec(text("DELETE FROM stock"))
            session.exec(text("DELETE FROM product"))
            session.commit()

    def transfer_stock(self, sku: str, from_warehouse: str, to_warehouse: str, quantity: int) -> None:
        '''
        Перемещает указанное количество товара с одного склада на другой.
        
        Args:
            sku (str): SKU товара
            from_warehouse (str): Склад-источник
            to_warehouse (str): Склад-назначение
            quantity (int): Количество для перемещения
            
        Raises:
            ValueError: Если количество отрицательное или равно нулю
            ValueError: Если склады совпадают
            ValueError: Если товара нет на складе-источнике
            ValueError: Если недостаточно товара на складе-источнике
        '''
        try:
            logging.info(
                f"Запрос на перемещение товара: SKU={sku}, "
                f"количество={quantity}, со склада={from_warehouse}, "
                f"на склад={to_warehouse}"
            )
            
            # Проверяем корректность входных данных
            if quantity <= 0:
                raise ValueError("Количество должно быть положительным числом")
                
            if from_warehouse == to_warehouse:
                raise ValueError("Склады источника и назначения не могут совпадать")
            
            # Используем существующий метод transfer
            self.transfer(sku, from_warehouse, to_warehouse, quantity)
            
            logging.info(
                f"Успешно перемещено {quantity} единиц товара {sku} "
                f"со склада {from_warehouse} на склад {to_warehouse}"
            )
            
        except Exception as e:
            logging.error(f"Ошибка при перемещении товара: {str(e)}")
            raise

    def get_product_sales_history(self, sku: str) -> List[dict]:
        '''Получить историю продаж конкретного товара.

        Args:
            sku (str): SKU товара

        Returns:
            List[dict]: Список словарей с информацией о продажах
        '''
        with Session(self.engine) as session:
            sales = session.exec(
                select(Sale)
                .where(Sale.sku == sku)
                .order_by(Sale.timestamp.desc())
            ).all()
            
            return [
                {
                    'id': sale.id,
                    'warehouse': sale.warehouse,
                    'quantity': sale.quantity,
                    'timestamp': sale.timestamp.isoformat()
                }
                for sale in sales
            ]


    def add_to_warehouse(self, sku: str, warehouse: str, quantity: int) -> None:
        '''
        Пополняет указанное количество товара на складе.
        
        Args:
            sku (str): SKU товара
            warehouse (str): Склад для пополнения
            quantity (int): Количество для добавления
            
        Raises:
            ValueError: Если количество отрицательное или равно нулю
            ValueError: Если товар не существует в базе данных
        '''
        try:
            logging.info(
                f"Запрос на пополнение товара: SKU={sku}, "
                f"количество={quantity}, склад={warehouse}"
            )
            
            # Проверяем корректность входных данных
            if quantity <= 0:
                raise ValueError("Количество должно быть положительным числом")
            
            with Session(self.engine) as session:
                # Проверяем существование товара
                product = session.get(Product, sku)
                if not product:
                    raise ValueError(f"Товар с SKU {sku} не найден в базе данных")
                
                # Получаем текущий остаток на складе
                stock = session.get(Stock, (sku, warehouse))
                
                if stock:
                    # Если остаток существует, увеличиваем его
                    stock.quantity += quantity
                    logging.info(f"Увеличен остаток на складе {warehouse}: было {stock.quantity - quantity}, стало {stock.quantity}")
                else:
                    # Если остатка нет, создаем новый
                    stock = Stock(sku=sku, warehouse=warehouse, quantity=quantity)
                    logging.info(f"Создан новый остаток на складе {warehouse}: {quantity}")
                
                session.add(stock)
                session.commit()
                
                # Запускаем задачу синхронизации Allegro по имени (без импорта)
                celery.send_task('app.services.allegro.sync_tasks.sync_allegro_stock_single_product', args=[sku])
                
                logging.info(
                    f"Успешно добавлено {quantity} единиц товара {sku} "
                    f"на склад {warehouse}"
                )
                
        except Exception as e:
            logging.error(f"Ошибка при пополнении товара: {str(e)}")
            raise

    def get_sales_statistics(self, skus: Optional[List[str]] = None) -> Dict[str, Dict[str, int]]:
        '''Получить статистику продаж за последние 15, 30 и 60 дней по SKU.
        
        Returns:
            Dict[str, Dict[str, int]]: Словарь в формате {sku: {'15d': кол-во, '30d': кол-во, '60d': кол-во}}
        '''
        now = datetime.utcnow()
        periods = {
            '15d': (now - timedelta(days=15)),
            '30d': (now - timedelta(days=30)),
            '60d': (now - timedelta(days=60))
        }
        
        result = {}
        
        with Session(self.engine) as session:
            # Базовый запрос
            base_query = select(Sale)
            if skus:
                base_query = base_query.where(Sale.sku.in_(skus))
            
            # Получаем все продажи за последние 60 дней
            sales = session.exec(
                base_query.where(Sale.timestamp >= periods['60d'])
            ).all()
            
            # Группируем продажи по SKU
            sales_by_sku = {}
            for sale in sales:
                if sale.sku not in sales_by_sku:
                    sales_by_sku[sale.sku] = []
                sales_by_sku[sale.sku].append(sale)
            
            # Подсчитываем количество для каждого периода
            for sku, sku_sales in sales_by_sku.items():
                result[sku] = {
                    '15d': sum(s.quantity for s in sku_sales if s.timestamp >= periods['15d']),
                    '30d': sum(s.quantity for s in sku_sales if s.timestamp >= periods['30d']),
                    '60d': sum(s.quantity for s in sku_sales if s.timestamp >= periods['60d'])
                }
        
        return result
from sqlmodel import SQLModel, Field, Session, create_engine, select
from sqlalchemy import Column, JSON, text
from typing import List, Optional, Tuple
import pandas as pd
from io import BytesIO
from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Query
from fastapi.responses import JSONResponse, StreamingResponse
from datetime import datetime, date
from enum import Enum
import logging
import openpyxl

from app.models.werehouse import Product, Stock, Sale
from app.core.config import settings

class Werehouses(str, Enum):
    """Перечисление доступных складов в системе.
    
    Attributes:
        A (str): Основной склад для хранения товаров
        B (str): Вторичный склад для хранения товаров
    """
    A = 'A'  # Основной склад
    B = 'B'  # Вторичный склад

def get_manager() -> 'InventoryManager':
    dsn = settings.SQLALCHEMY_DATABASE_URI.unicode_string()
    return InventoryManager(dsn)

class InventoryManager:
    def __init__(self, dsn: str):
        '''Менеджер инвентаря через PostgreSQL / SQLModel.'''
        self.engine = create_engine(dsn)
        # SQLModel.metadata.create_all(self.engine)

    def restock(self, sku: str, warehouse: str, quantity: int):
        '''Пополнение остатков на указанном складе.'''
        if quantity <= 0:
            raise ValueError('Количество пополнения должно быть положительным.')
        with Session(self.engine) as session:
            key = (sku, warehouse)
            stock = session.get(Stock, key)
            if not stock:
                stock = Stock(sku=sku, warehouse=warehouse, quantity=quantity)
            else:
                stock.quantity += quantity
            session.add(stock)
            session.commit()

    def transfer(self, sku: str, quantity: int, from_warehouse: str, to_warehouse: str):
        '''Перемещение количества между любыми складами.'''
        logging.info(f"Начало перемещения: SKU={sku}, количество={quantity}, со склада={from_warehouse}, на склад={to_warehouse}")
        
        if quantity <= 0:
            logging.error(f"Ошибка: количество должно быть положительным, получено {quantity}")
            raise ValueError('Количество перевода должно быть положительным.')
            
        if from_warehouse == to_warehouse:
            logging.error(f"Ошибка: склады совпадают: {from_warehouse}")
            raise ValueError('Склады отправления и назначения должны различаться.')
            
        with Session(self.engine) as session:
            # Проверяем наличие товара на складе-источнике
            src = session.get(Stock, (sku, from_warehouse))
            logging.info(f"Проверка склада-источника: найден={src is not None}, "
                        f"текущий остаток={src.quantity if src else 0}")
            
            if not src:
                logging.error(f"Ошибка: товар {sku} не найден на складе {from_warehouse}")
                raise ValueError(f'Недостаточно на складе {from_warehouse}.')
            
            if src.quantity < quantity:
                logging.error(f"Ошибка: недостаточно товара. Требуется {quantity}, в наличии {src.quantity}")
                raise ValueError(f'Недостаточно на складе {from_warehouse}. Требуется {quantity}, в наличии {src.quantity}')
            
            # Проверяем склад назначения
            dest = session.get(Stock, (sku, to_warehouse))
            logging.info(f"Проверка склада-назначения: найден={dest is not None}, "
                        f"текущий остаток={dest.quantity if dest else 0}")
            
            # Уменьшаем количество на складе-источнике
            src.quantity -= quantity
            logging.info(f"Уменьшен остаток на складе {from_warehouse}: было {src.quantity + quantity}, стало {src.quantity}")
            
            # Увеличиваем количество на складе-назначения
            if not dest:
                dest = Stock(sku=sku, warehouse=to_warehouse, quantity=quantity)
                logging.info(f"Создан новый остаток на складе {to_warehouse}: {quantity}")
            else:
                dest.quantity += quantity
                logging.info(f"Увеличен остаток на складе {to_warehouse}: было {dest.quantity - quantity}, стало {dest.quantity}")
            
            try:
                session.add(src)
                session.add(dest)
                session.commit()
                logging.info("Перемещение успешно завершено")
            except Exception as e:
                logging.error(f"Ошибка при сохранении изменений: {str(e)}")
                session.rollback()
                raise

    def remove_from_warehouse(self, sku: str, warehouse: str, quantity: int):
        '''Списание количества с указанного склада + логирование.'''
        if quantity <= 0:
            raise ValueError('Количество списания должно быть положительным.')
        with Session(self.engine) as session:
            stock = session.get(Stock, (sku, warehouse))
            if not stock or stock.quantity < quantity:
                raise ValueError(f'Недостаточно на складе {warehouse}.')
            stock.quantity -= quantity
            session.add(stock)
            # логирование списания
            sale = Sale(sku=sku, warehouse=warehouse, quantity=quantity)
            session.add(sale)
            session.commit()

    def remove_one(self, sku: str, warehouse: str):
        '''Списание одной единицы с указанного склада.'''
        self.remove_from_warehouse(sku, warehouse, 1)

    def import_incoming_from_excel_depricated(
        self,
        file_bytes: bytes,
        warehouse: str,
        sku_col: str = 'sku',
        qty_col: str = 'quantity',
        ean_col: str = 'ean',
        name_col: str = 'name',
        image_col: str = 'Foto',
        header: int = 0
    ):
        '''Импорт прихода из XLSX (bytes): SKU, количество, EAN, название и изображение.'''
        logging.info(f"Начало импорта Excel файла. Размер: {len(file_bytes)} байт")
        
        # Загружаем Excel файл
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        ws = wb.active
        logging.info(f"Excel файл загружен. Активный лист: {ws.title}")
        
        # Получаем заголовки и их индексы
        headers = [str(cell.value) for cell in ws[header + 1]]
        logging.info(f"Найдены заголовки: {headers}")
        
        # Находим индексы нужных колонок
        try:
            col_indices = {
                'sku': headers.index(sku_col),
                'qty': headers.index(qty_col),
                'ean': headers.index(ean_col),
                'name': headers.index(name_col)
            }
            logging.info(f"Индексы колонок: {col_indices}")
        except ValueError as e:
            logging.error(f"Ошибка при поиске колонок: {str(e)}")
            raise ValueError(f"Не найдена одна из обязательных колонок. Требуются: {sku_col}, {qty_col}, {ean_col}, {name_col}")

        # Создаем словарь изображений по номерам строк
        images_by_row = {}
        logging.info(f"Всего изображений в документе: {len(ws._images)}")
        
        for img in ws._images:
            coords = img.anchor._from
            row = coords.row + 1  # Преобразуем в 1-based индекс
            if row > (header + 1):  # Пропускаем строку с заголовками
                try:
                    logging.info(f"Обработка изображения: тип={type(img)}, тип ref={type(img.ref)}")
                    # Получаем байты изображения
                    if hasattr(img.ref, 'getvalue'):
                        img_bytes = img.ref.getvalue()  # Используем getvalue() для BytesIO
                        images_by_row[row] = img_bytes
                        logging.info(f"Найдено изображение для строки {row}. Размер: {len(img_bytes)} байт")
                    elif hasattr(img.ref, '_data'):
                        img_bytes = img.ref._data()
                        images_by_row[row] = img_bytes
                        logging.info(f"Найдено изображение (метод _data) для строки {row}. Размер: {len(img_bytes)} байт")
                    elif hasattr(img.ref, 'content'):
                        img_bytes = img.ref.content
                        images_by_row[row] = img_bytes
                        logging.info(f"Найдено изображение (через content) для строки {row}. Размер: {len(img_bytes)} байт")
                    else:
                        available_attrs = dir(img.ref)
                        logging.error(f"Не найден метод для извлечения данных изображения. Доступные атрибуты: {available_attrs}")
                except Exception as e:
                    logging.error(f"Ошибка при извлечении изображения из строки {row}: {str(e)}")

        # Обрабатываем каждую строку данных
        for row_idx, row in enumerate(ws.iter_rows(min_row=header + 2), start=header + 2):
            try:
                sku = str(row[col_indices['sku']].value)
                qty = int(row[col_indices['qty']].value)
                ean = str(row[col_indices['ean']].value)
                name = str(row[col_indices['name']].value)
                
                logging.info(f"Обработка строки {row_idx}. SKU: {sku}, Name: {name}")
                
                # Получаем изображение для текущей строки
                image_data = images_by_row.get(row_idx)
                if image_data:
                    logging.info(f"Найдено изображение для SKU {sku}")
                
                with Session(self.engine) as session:
                    product = session.get(Product, sku)
                    if not product:
                        product = Product(sku=sku, eans=[ean], name=name, image=image_data)
                        session.add(product)
                        logging.info(f"Создан новый продукт: {sku}")
                    else:
                        # Обновляем все поля продукта, кроме SKU
                        if ean not in product.eans:
                            product.eans.append(ean)
                        product.name = name
                        if image_data:
                            product.image = image_data
                            logging.info(f"Обновлено изображение для продукта: {sku}")
                        session.add(product)
                    session.commit()
                
                # Добавляем или обновляем остатки на складе
                self.restock(sku, warehouse, qty)
            except Exception as e:
                logging.error(f"Ошибка при обработке строки {row_idx}: {str(e)}")
                continue
        
        logging.info("Импорт Excel файла завершен")

    def import_incoming_from_excel(
        self,
        file_bytes: bytes,
        warehouse: str,
        sku_col: str = 'sku',
        qty_col: str = 'quantity',
        ean_col: str = 'ean',
        name_col: str = 'name',
        image_col: str = 'Foto',
        header: int = 0
    ):
        '''Импорт прихода из XLSX (bytes): SKU, количество, EAN, название и изображение.'''
        logging.info("Начало импорта Excel файла")
        
        # Загружаем Excel файл
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        ws = wb.active
        
        # Получаем заголовки и их индексы
        headers = [str(cell.value) for cell in ws[header + 1]]
        
        # Находим индексы нужных колонок
        try:
            col_indices = {
                'sku': headers.index(sku_col),
                'qty': headers.index(qty_col),
                'ean': headers.index(ean_col),
                'name': headers.index(name_col)
            }
        except ValueError as e:
            raise ValueError(f"Не найдена одна из обязательных колонок. Требуются: {sku_col}, {qty_col}, {ean_col}, {name_col}")

        # Создаем словарь изображений по номерам строк
        images_by_row = {}
        
        for img in ws._images:
            coords = img.anchor._from
            row = coords.row + 1
            if row > (header + 1):
                try:
                    if hasattr(img.ref, 'getvalue'):
                        images_by_row[row] = img.ref.getvalue()
                    elif hasattr(img.ref, '_data'):
                        images_by_row[row] = img.ref._data()
                    elif hasattr(img.ref, 'content'):
                        images_by_row[row] = img.ref.content
                except Exception:
                    continue

        # Обрабатываем каждую строку данных
        for row_idx, row in enumerate(ws.iter_rows(min_row=header + 2), start=header + 2):
            try:
                sku = str(row[col_indices['sku']].value)
                
                # Пропускаем строку если SKU пустой
                if not sku or sku.lower() == 'none':
                    continue
                    
                qty = int(row[col_indices['qty']].value)
                ean = str(row[col_indices['ean']].value)
                name = str(row[col_indices['name']].value)
                image_data = images_by_row.get(row_idx)
                
                with Session(self.engine) as session:
                    product = session.get(Product, sku)
                    if not product:
                        product = Product(sku=sku, eans=[ean], name=name, image=image_data)
                        session.add(product)
                    else:
                        if ean not in product.eans:
                            product.eans.append(ean)
                        product.name = name
                        if image_data:
                            product.image = image_data
                        session.add(product)
                    session.commit()
                
                # Добавляем или обновляем остатки на складе
                self.restock(sku, warehouse, qty)
            except Exception:
                continue
        
        logging.info("Импорт Excel файла завершен")

    def import_transfer_from_excel(
        self,
        file_bytes: bytes,
        from_warehouse: str,
        to_warehouse: str,
        sku_col: str = 'sku',
        qty_col: str = 'кол-во',
        header: int = 1
    ):
        '''Импорт перемещения из XLSX (bytes): SKU и количество.'''
        df = pd.read_excel(BytesIO(file_bytes), header=header)
        for _, row in df.iterrows():
            try:
                # Пропускаем строку если SKU или количество - nan
                if pd.isna(row[sku_col]) or pd.isna(row[qty_col]):
                    continue
                    
                sku = str(row[sku_col])
                # Пропускаем если SKU пустой
                if not sku or sku.lower() == 'none':
                    continue
                    
                qty = int(row[qty_col])
                self.transfer(sku, qty, from_warehouse, to_warehouse)
            except KeyError as e:
                raise ValueError(f'Колонка не найдена: {e}. Ожидаются колонки: {sku_col}, {qty_col}')
            except ValueError as e:
                raise ValueError(f'Ошибка в данных: {str(e)}. SKU: {sku}, количество: {qty}')

    def get_stock_report(self) -> Tuple[pd.DataFrame, List[bytes]]:
        """
        Возвращает кортеж:
        - DataFrame с колонками sku, name, eans, warehouse_*, total (без image)
        - Список байтов изображений в том же порядке, что и строки в DataFrame
        """
        with Session(self.engine) as session:
            products = session.exec(select(Product)).all()
            stocks   = session.exec(select(Stock)).all()

        if not products:
            empty_df = pd.DataFrame(columns=[
                'sku', 'name', 'eans', 'total'
            ] + [f'warehouse_{w.value}' for w in Werehouses])
            return empty_df, []

        # накапливаем строки
        rows = []
        images = []
        # создаём словарь stock_data
        stock_data = {s.sku: {w.value: 0 for w in Werehouses} for s in stocks}
        for s in stocks:
            stock_data.setdefault(s.sku, {w.value: 0 for w in Werehouses})
            stock_data[s.sku][s.warehouse] = s.quantity

        for prod in products:
            # Собираем общие данные (без image)
            row = {
                'sku': prod.sku,
                'name': prod.name,
                'eans': ', '.join(prod.eans) if prod.eans else '',
            }
            total = 0
            for wh in Werehouses:
                qty = stock_data.get(prod.sku, {}).get(wh.value, 0)
                row[f'warehouse_{wh.value}'] = qty
                total += qty
            row['total'] = total

            rows.append(row)
            # А картинку — собираем параллельно в отдельный список
            images.append(prod.image or b'')

        # Теперь df без колонки image
        df = pd.DataFrame(rows)
        return df, images

    def get_sales_report(self, start_date: date, end_date: date, sku: Optional[str] = None) -> pd.DataFrame:
        '''Получить DataFrame логов продаж за период по SKU.'''
        with Session(self.engine) as session:
            stmt = select(Sale).where(Sale.timestamp >= datetime.combine(start_date, datetime.min.time()),
                                     Sale.timestamp <= datetime.combine(end_date, datetime.max.time()))
            if sku:
                stmt = stmt.where(Sale.sku == sku)
            sales = session.exec(stmt).all()
        rows = [{'sku': s.sku, 'warehouse': s.warehouse, 'quantity': s.quantity, 'timestamp': s.timestamp} for s in sales]
        return pd.DataFrame(rows)

    def get_stock_by_sku(self, sku: str) -> dict:
        '''Получить остатки по конкретному SKU на всех складах.'''
        with Session(self.engine) as session:
            stocks = session.exec(select(Stock).where(Stock.sku == sku)).all()
            return {stock.warehouse: stock.quantity for stock in stocks}

    def clear_all_data(self):
        '''Удаляет все продукты, остатки и логи продаж из базы данных.'''
        with Session(self.engine) as session:
            session.exec(text("DELETE FROM sale"))
            session.exec(text("DELETE FROM stock"))
            session.exec(text("DELETE FROM product"))
            session.commit()

    def transfer_stock(self, sku: str, from_warehouse: str, to_warehouse: str, quantity: int) -> None:
        '''
        Перемещает указанное количество товара с одного склада на другой.
        
        Args:
            sku (str): SKU товара
            from_warehouse (str): Склад-источник
            to_warehouse (str): Склад-назначение
            quantity (int): Количество для перемещения
            
        Raises:
            ValueError: Если количество отрицательное или равно нулю
            ValueError: Если склады совпадают
            ValueError: Если товара нет на складе-источнике
            ValueError: Если недостаточно товара на складе-источнике
        '''
        try:
            logging.info(
                f"Запрос на перемещение товара: SKU={sku}, "
                f"количество={quantity}, со склада={from_warehouse}, "
                f"на склад={to_warehouse}"
            )
            
            # Проверяем корректность входных данных
            if quantity <= 0:
                raise ValueError("Количество должно быть положительным числом")
                
            if from_warehouse == to_warehouse:
                raise ValueError("Склад источника и назначения не могут совпадать")
            
            # Используем существующий метод transfer
            self.transfer(sku, quantity, from_warehouse, to_warehouse)
            
            logging.info(
                f"Успешно перемещено {quantity} единиц товара {sku} "
                f"со склада {from_warehouse} на склад {to_warehouse}"
            )
            
        except Exception as e:
            logging.error(f"Ошибка при перемещении товара: {str(e)}")
            raise

    def get_product_sales_history(self, sku: str) -> List[dict]:
        '''Получить историю продаж конкретного товара.

        Args:
            sku (str): SKU товара

        Returns:
            List[dict]: Список словарей с информацией о продажах
        '''
        with Session(self.engine) as session:
            sales = session.exec(
                select(Sale)
                .where(Sale.sku == sku)
                .order_by(Sale.timestamp.desc())
            ).all()
            
            return [
                {
                    'id': sale.id,
                    'warehouse': sale.warehouse,
                    'quantity': sale.quantity,
                    'timestamp': sale.timestamp.isoformat()
                }
                for sale in sales
            ]

    def add_to_warehouse(self, sku: str, warehouse: str, quantity: int) -> None:
        '''
        Пополняет указанное количество товара на складе.
        
        Args:
            sku (str): SKU товара
            warehouse (str): Склад для пополнения
            quantity (int): Количество для добавления
            
        Raises:
            ValueError: Если количество отрицательное или равно нулю
            ValueError: Если товар не существует в базе данных
        '''
        try:
            logging.info(
                f"Запрос на пополнение товара: SKU={sku}, "
                f"количество={quantity}, склад={warehouse}"
            )
            
            # Проверяем корректность входных данных
            if quantity <= 0:
                raise ValueError("Количество должно быть положительным числом")
            
            with Session(self.engine) as session:
                # Проверяем существование товара
                product = session.get(Product, sku)
                if not product:
                    raise ValueError(f"Товар с SKU {sku} не найден в базе данных")
                
                # Получаем текущий остаток на складе
                stock = session.get(Stock, (sku, warehouse))
                
                if stock:
                    # Если остаток существует, увеличиваем его
                    stock.quantity += quantity
                    logging.info(f"Увеличен остаток на складе {warehouse}: было {stock.quantity - quantity}, стало {stock.quantity}")
                else:
                    # Если остатка нет, создаем новый
                    stock = Stock(sku=sku, warehouse=warehouse, quantity=quantity)
                    logging.info(f"Создан новый остаток на складе {warehouse}: {quantity}")
                
                session.add(stock)
                session.commit()
                
                logging.info(
                    f"Успешно добавлено {quantity} единиц товара {sku} "
                    f"на склад {warehouse}"
                )
                
        except Exception as e:
            logging.error(f"Ошибка при пополнении товара: {str(e)}")
            raise
import logging
from typing import List, Optional
import pandas as pd
from io import BytesIO
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, UploadFile, File, Query, Request, Form
from fastapi.responses import JSONResponse, StreamingResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlmodel.ext.asyncio.session import AsyncSession
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from app.api import deps
from app.services.allegro import tokens as allegro_tokens
from app.services.allegro.data_access import get_tokens_list, insert_token, delete_token, get_token_by_id
from app.services.allegro.pydantic_models import TokenOfAllegro
from app.services.allegro.pydantic_models import InitializeAuth
from app.models.user import User as UserModel
from app.services.werehouse import manager
from app.services.werehouse.manager import Werehouses
from pydantic import BaseModel

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

# Добавляем модель для запроса перемещения
class TransferItem(BaseModel):
    sku: str
    from_warehouse: str
    to_warehouse: str
    quantity: int

class RemoveItem(BaseModel):
    sku: str
    warehouse: str
    quantity: int

class AddItem(BaseModel):
    sku: str
    warehouse: str
    quantity: int

@router.get("/inventory", response_class=HTMLResponse)
async def inventory_page(request: Request, current_user: UserModel = Depends(deps.get_current_user)):
    return templates.TemplateResponse("inventory.html", {
        "request": request,
        "user": current_user,
        "warehouses": [w.value for w in Werehouses]
    })

@router.get("/operations", response_class=HTMLResponse)
async def operations_page(request: Request, current_user: UserModel = Depends(deps.get_current_user)):
    return templates.TemplateResponse("operations.html", {
        "request": request,
        "user": current_user
    })

@router.post('/incoming/', summary='Импорт прихода на склад')
async def upload_incoming(
    file: UploadFile = File(...),
    warehouse: Werehouses = Query(Werehouses.A, description="Склад для импорта товаров"),
    sku_col: str = 'sku',
    qty_col: str = 'кол-во',
    ean_col: str = 'EAN/UPC',
    name_col: str = 'Name',
    image_col: str = 'Foto',
    header: int = 1,
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='Неверный формат файла. Ожидается XLS/XLSX.')
    content = await file.read()
    try:
        manager.import_incoming_from_excel(content, warehouse.value, sku_col, qty_col, ean_col, name_col, image_col, header)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return JSONResponse({'status': 'success', 'file': file.filename, 'warehouse': warehouse.value})

@router.post('/transfer/', summary='Импорт перемещения между складами')
async def upload_transfer(
    file: UploadFile = File(...),
    from_warehouse: str = Form(..., description="Склад-источник"),
    to_warehouse: str = Form(..., description="Склад-назначение"),
    sku_col: str = Form('sku'),
    qty_col: str = Form('кол-во'),
    header: int = Form(1),
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='Неверный формат файла. Ожидается XLS/XLSX.')
    
    # Проверяем и конвертируем значения складов
    try:
        from_wh = Werehouses(from_warehouse)
        to_wh = Werehouses(to_warehouse)
    except ValueError:
        raise HTTPException(status_code=400, detail='Неверное значение склада')
    
    content = await file.read()
    try:
        manager.import_transfer_from_excel(content, from_wh.value, to_wh.value, sku_col, qty_col, header)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return JSONResponse({'status': 'success', 'file': file.filename, 'from': from_wh.value, 'to': to_wh.value})

@router.get(
    "/export/stock/",
    summary="Экспорт остатков",
    response_class=StreamingResponse,
)
async def export_stock(
    skus: Optional[List[str]] = Query(None, description="Список SKU для экспорта"),
    manager: manager.InventoryManager = Depends(manager.get_manager),
):
    """Возвращает XLSX-файл с остатками и картинками 100×100px."""
    # получаем DF без image и параллельный список bytes
    df, images = manager.get_stock_report()

    # фильтрация, если skus заданы
    if skus:
        mask = df["sku"].isin(skus)
        df = df[mask].reset_index(drop=True)
        images = [img for keep, img in zip(mask, images) if keep]

    # переименования для читабельности
    column_renames = {
        "sku": "SKU",
        "eans": "EAN коды",
        "name": "Наименование",
        "total": "Общий остаток"
    }
    # добавляем названия складов
    for wh in Werehouses:
        column_renames[f"warehouse_{wh.value}"] = f"Склад {wh.value}"
    
    df = df.rename(columns=column_renames)

    # вставляем пустую колонку-«держатель» под изображения
    df.insert(1, "Изображение", "")

    # порядок колонок
    cols = ["SKU", "Изображение", "Наименование", "EAN коды"]
    cols.extend([f"Склад {wh.value}" for wh in Werehouses])
    cols.append("Общий остаток")
    df = df[cols]

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # пишем данные (картинки пока пустые)
        df.to_excel(writer, index=False, sheet_name="Остатки")
        ws = writer.sheets["Остатки"]

        # находим букву колонки «Изображение»
        img_col_idx    = df.columns.get_loc("Изображение") + 1
        img_col_letter = get_column_letter(img_col_idx)

        # 1) форматируем заголовки и ширину
        for idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.font = cell.font.copy(bold=True)
            if col == "Изображение":
                ws.column_dimensions[img_col_letter].width = 18  # ~100px
            else:
                max_len = max(df[col].astype(str).map(len).max(), len(col))
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)

        # 2) вставка картинок 100×100 и высота строк
        for row_idx, img_bytes in enumerate(images, start=2):
            if not img_bytes:
                continue
            try:
                img = XLImage(BytesIO(img_bytes))
                img.width  = 100
                img.height = 100
                ws.row_dimensions[row_idx].height = 75  # под ~100px

                ws.add_image(img, f"{img_col_letter}{row_idx}")
            except Exception as e:
                logging.error(f"Не удалось вставить изображение в строке {row_idx}: {e}")

    buf.seek(0)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"stock_report_{ts}.xlsx"
    headers = {
        "Content-Disposition": (
            f'attachment; filename="{filename}"; '
            f"filename*=UTF-8''{filename}"
        )
    }
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers
    )

@router.get('/export/stock/no-images/', summary='Экспорт остатков без изображений') 
async def export_stock_no_images(manager: manager.InventoryManager = Depends(manager.get_manager)):
    '''Возвращает XLSX-файл с остатками по складам и общим, без изображений.'''
    df, _ = manager.get_stock_report()  # Игнорируем список изображений
    
    # Русские названия столбцов
    column_renames = {
        "sku": "SKU",
        "eans": "EAN коды",
        "name": "Наименование",
        "total": "Общий остаток"
    }
    # добавляем названия складов
    for wh in Werehouses:
        column_renames[f"warehouse_{wh.value}"] = f"Склад {wh.value}"
    
    df = df.rename(columns=column_renames)
    
    # порядок колонок
    cols = ["SKU", "Наименование", "EAN коды"]
    cols.extend([f"Склад {wh.value}" for wh in Werehouses])
    cols.append("Общий остаток")
    df = df[cols]
    
    # Создаем Excel-файл
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Остатки')
        worksheet = writer.sheets['Остатки']
        
        # Форматируем заголовки и ширины
        for idx, col in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=idx)
            cell.font = cell.font.copy(bold=True)
            
            # Устанавливаем ширину колонок
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            worksheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)
    
    buffer.seek(0)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"stock_report_no_images_{ts}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
    }
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )

@router.get('/sales/report/', summary='Отчёт по продажам')
async def sales_report(
    start_date: date = Query(..., description='Начальная дата в формате YYYY-MM-DD'),
    end_date: date = Query(..., description='Конечная дата в формате YYYY-MM-DD'),
    sku: Optional[str] = Query(None, description='Фильтр по SKU'),
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Возвращает XLSX-файл с логами продаж за указанный период.'''
    df = manager.get_sales_report(start_date, end_date, sku)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sales')
    buffer.seek(0)
    filename = f"sales_{start_date}_{end_date}" + (f"_{sku}.xlsx" if sku else '.xlsx')
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)

@router.get('/stock/{sku}', summary='Проверка остатков по SKU')
async def check_stock(
    sku: str,
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Возвращает остатки по конкретному SKU на всех складах.'''
    stocks = manager.get_stock_by_sku(sku)
    return JSONResponse(content=stocks)

@router.delete('/clear-all/', summary='Удалить все данные')
async def clear_all_data(
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Удаляет все продукты, остатки и логи продаж из базы данных.'''
    manager.clear_all_data()
    return JSONResponse(content={"message": "Все данные успешно удалены"})

@router.post('/transfer-item/', summary='Перемещение товара между складами')
async def transfer_item(
    transfer: TransferItem,
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Перемещает указанное количество товара с одного склада на другой.'''
    try:
        # Проверяем и конвертируем значения складов
        from_wh = Werehouses(transfer.from_warehouse)
        to_wh = Werehouses(transfer.to_warehouse)
        
        if from_wh == to_wh:
            raise HTTPException(
                status_code=400,
                detail='Склады отправления и назначения должны быть разными'
            )
        
        # Проверяем наличие достаточного количества товара на складе-источнике
        stocks = manager.get_stock_by_sku(transfer.sku)
        if not stocks or transfer.from_warehouse not in stocks:
            raise HTTPException(
                status_code=404,
                detail='Товар не найден на складе-источнике'
            )
        
        if stocks[transfer.from_warehouse] < transfer.quantity:
            raise HTTPException(
                status_code=400,
                detail='Недостаточное количество товара на складе-источнике'
            )
        
        # Выполняем перемещение
        manager.transfer_stock(
            transfer.sku,
            from_wh.value,
            to_wh.value,
            transfer.quantity
        )
        
        return JSONResponse({
            'status': 'success',
            'message': f'Успешно перемещено {transfer.quantity} шт. со склада {from_wh.value} на склад {to_wh.value}'
        })
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail='Неверное значение склада')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get('/product/{sku}/sales/', summary='История продаж товара')
async def get_product_sales_history(
    sku: str,
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Возвращает историю продаж конкретного товара.'''
    try:
        sales = manager.get_product_sales_history(sku)
        return JSONResponse(content=sales)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post('/remove/', summary='Списание товара со склада')
async def remove_from_stock(
    item: RemoveItem,
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Списывает указанное количество товара с определенного склада.'''
    try:
        # Проверяем корректность склада
        try:
            warehouse = Werehouses(item.warehouse)
        except ValueError:
            raise HTTPException(status_code=400, detail='Неверное значение склада')
        
        # Проверяем наличие достаточного количества товара
        stocks = manager.get_stock_by_sku(item.sku)
        if not stocks or item.warehouse not in stocks:
            raise HTTPException(
                status_code=404,
                detail='Товар не найден на складе'
            )
        
        if stocks[item.warehouse] < item.quantity:
            raise HTTPException(
                status_code=400,
                detail='Недостаточное количество товара на складе'
            )
        
        # Выполняем списание
        manager.remove_from_warehouse(
            item.sku,
            warehouse.value,
            item.quantity
        )
        
        return JSONResponse({
            'status': 'success',
            'message': f'Успешно списано {item.quantity} шт. со склада {warehouse.value}'
        })
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post('/add/', summary='Пополнение товара на складе')
async def add_to_stock(
    item: AddItem,
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Пополняет указанное количество товара на определенном складе.'''
    try:
        # Проверяем корректность склада
        try:
            warehouse = Werehouses(item.warehouse)
        except ValueError:
            raise HTTPException(status_code=400, detail='Неверное значение склада')
        
        # Проверяем существование товара
        stocks = manager.get_stock_by_sku(item.sku)
        if not stocks:
            raise HTTPException(
                status_code=404,
                detail='Товар не найден'
            )
        
        # Выполняем пополнение
        manager.add_to_warehouse(
            item.sku,
            warehouse.value,
            item.quantity
        )
        
        return JSONResponse({
            'status': 'success',
            'message': f'Успешно добавлено {item.quantity} шт. на склад {warehouse.value}'
        })
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

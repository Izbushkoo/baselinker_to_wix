import logging
import mimetypes
from typing import List, Optional
from sqlmodel import select, update
from app.models.allegro_order import AllegroOrder
import pandas as pd
from io import BytesIO
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, UploadFile, File, Query, Request, Form
from fastapi.responses import JSONResponse, StreamingResponse, HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlmodel.ext.asyncio.session import AsyncSession
import openpyxl
from openpyxl.utils import get_column_letter
from app.models.user import User
from openpyxl.drawing.image import Image as XLImage
from app.api import deps
from app.services.allegro import tokens as allegro_tokens
from app.services.allegro.data_access import get_tokens_list, insert_token, delete_token, get_token_by_id
from app.services.allegro.pydantic_models import TokenOfAllegro
from app.services.allegro.pydantic_models import InitializeAuth
from app.models.user import User as UserModel
from app.services.warehouse import manager
from app.services.warehouse.manager import Warehouses
from pydantic import BaseModel, Field
from openpyxl.styles import Font, Alignment
from urllib.parse import quote
from app.services.operations_service import OperationsService, get_operations_service
from app.models.operations import OperationType

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

# Добавляем модель для запроса перемещения
class TransferItem(BaseModel):
    sku: str
    from_warehouse: str
    to_warehouse: str
    quantity: int

class RemoveItem(BaseModel):
    sku: str
    warehouse: str
    quantity: int

class AddItem(BaseModel):
    sku: str
    warehouse: str
    quantity: int

class SaleFromOrder(BaseModel):
    order_id: str
    line_items: List[dict]
    warehouse: str = Field(default=Warehouses.A.value, description="Склад для списания товаров")

@router.get("/operations", response_class=HTMLResponse)
async def operations_page(request: Request, current_user: User = Depends(deps.get_current_user_optional)):
    logging.info(f"Operations page requested by user: {current_user}")
    if not current_user:
        return RedirectResponse(url=f"/login?next=/operations", status_code=302)
    return templates.TemplateResponse("operations.html", {
        "request": request,
        "user": current_user
    })

@router.post('/incoming/', summary='Импорт прихода на склад')
async def upload_incoming(
    file: UploadFile = File(...),
    warehouse: Warehouses = Query(..., description="Склад для импорта товаров"),
    sku_col: str = 'sku',
    qty_col: str = 'Кол-во',
    ean_col: str = 'EAN',
    name_col: str = 'Name',
    image_col: str = 'Foto',
    header: int = 0,
    manager: manager.InventoryManager = Depends(manager.get_manager),
    current_user: User = Depends(deps.get_current_user_optional)
):
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='Неверный формат файла. Ожидается XLS/XLSX.')
    content = await file.read()
    try:
        processed_products = manager.import_incoming_from_excel(content, warehouse.value, sku_col, qty_col, ean_col, name_col, image_col, header)
        
        # Создаем запись операции
        operations_service = get_operations_service()
        operations_service.create_file_operation(
            operation_type=OperationType.STOCK_IN_FILE,
            warehouse_id=warehouse.value,
            user_email=current_user.email,
            file_name=file.filename,
            products=processed_products
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return JSONResponse({'status': 'success', 'file': file.filename, 'warehouse': warehouse.value})

@router.post('/transfer/', summary='Импорт перемещения между складами')
async def upload_transfer(
    file: UploadFile = File(...),
    from_warehouse: str = Form(..., description="Склад-источник"),
    to_warehouse: str = Form(..., description="Склад-назначение"),
    sku_col: str = Form('sku'),
    qty_col: str = Form('Кол-во'),
    manager: manager.InventoryManager = Depends(manager.get_manager),
    current_user: User = Depends(deps.get_current_user_optional)
):
    if not file.filename.lower().endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail='Неверный формат файла. Ожидается XLS/XLSX.')
    
    # Проверяем и конвертируем значения складов
    try:
        from_wh = Warehouses(from_warehouse)
        to_wh = Warehouses(to_warehouse)
    except ValueError:
        raise HTTPException(status_code=400, detail='Неверное значение склада')
    
    content = await file.read()
    try:
        errors_df, processed_products = manager.import_transfer_from_excel(content, from_wh.value, to_wh.value, sku_col, qty_col)
        
        # Создаем запись операции
        operations_service = get_operations_service()
        operations_service.create_file_operation(
            operation_type=OperationType.TRANSFER_FILE,
            warehouse_id=from_wh.value,
            user_email=current_user.email,
            file_name=file.filename,
            target_warehouse_id=to_wh.value,
            products=processed_products
        )
        
        # Если есть ошибки, возвращаем файл с ошибками
        if not errors_df.empty:
            # Удаляем колонку с изображениями, если она есть
            if 'Foto' in errors_df.columns:
                errors_df = errors_df.drop('Foto', axis=1)
            
            # Создаем буфер для Excel файла
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Сохраняем данные
                errors_df.to_excel(writer, index=False)
                
                # Получаем рабочий лист для форматирования
                worksheet = writer.sheets['Sheet1']
                
                # Устанавливаем ширину колонок
                for idx, col in enumerate(errors_df.columns):
                    max_length = max(
                        errors_df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 2, 50)
            
            output.seek(0)
            
            # Формируем имя файла с ошибками и кодируем его
            filename = f"errors_{file.filename}"
            encoded_filename = quote(filename)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={encoded_filename}"
                }
            )
            
        return JSONResponse({'status': 'success', 'file': file.filename, 'from': from_wh.value, 'to': to_wh.value})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get('/stock/{sku}', summary='Проверка остатков по SKU')
async def check_stock(
    sku: str,
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Возвращает остатки по конкретному SKU на всех складах.'''
    stocks = manager.get_stock_by_sku(sku)
    return JSONResponse(content=stocks)

@router.delete('/clear-all/', summary='Удалить все данные')
async def clear_all_data(
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Удаляет все продукты, остатки и логи продаж из базы данных.'''
    manager.clear_all_data()
    return JSONResponse(content={"message": "Все данные успешно удалены"})

@router.post('/transfer-item/', summary='Перемещение товара между складами')
async def transfer_item(
    transfer: TransferItem,
    manager: manager.InventoryManager = Depends(manager.get_manager),
    current_user: User = Depends(deps.get_current_user_optional)
):
    '''Перемещает указанное количество товара с одного склада на другой.'''
    try:
        # Проверяем и конвертируем значения складов
        from_wh = Warehouses(transfer.from_warehouse)
        to_wh = Warehouses(transfer.to_warehouse)
        
        if from_wh == to_wh:
            raise HTTPException(
                status_code=400,
                detail='Склады отправления и назначения должны быть разными'
            )
        
        # Проверяем наличие достаточного количества товара на складе-источнике
        stocks = manager.get_stock_by_sku(transfer.sku)
        if not stocks or transfer.from_warehouse not in stocks:
            raise HTTPException(
                status_code=404,
                detail='Товар не найден на складе-источнике'
            )
        
        if stocks[transfer.from_warehouse] < transfer.quantity:
            raise HTTPException(
                status_code=400,
                detail='Недостаточное количество товара на складе-источнике'
            )
        
        # Выполняем перемещение
        manager.transfer_stock(
            transfer.sku,
            from_wh.value,
            to_wh.value,
            transfer.quantity
        )
        
        # Создаем запись операции
        operations_service = get_operations_service()
        operations_service.create_single_operation(
            operation_type=OperationType.TRANSFER,
            warehouse_id=from_wh.value,
            sku=transfer.sku,
            quantity=transfer.quantity,
            user_email=current_user.email,
            target_warehouse_id=to_wh.value,
        )
        
        return JSONResponse({
            'status': 'success',
            'message': f'Успешно перемещено {transfer.quantity} шт. со склада {from_wh.value} на склад {to_wh.value}'
        })
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail='Неверное значение склада')
    except Exception as e:
        logging.error(f"Ошибка при перемещении товара: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post('/remove/', summary='Списание товара со склада')
async def remove_from_stock(
    item: RemoveItem,
    manager: manager.InventoryManager = Depends(manager.get_manager),
    current_user: User = Depends(deps.get_current_user_optional)
):
    """Списывает указанное количество товара с определенного склада."""
    try:
        # Проверяем корректность склада
        try:
            warehouse = Warehouses(item.warehouse)
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail='Неверное значение склада'
            )
        
        # Проверяем наличие достаточного количества товара
        stocks = manager.get_stock_by_sku(item.sku)
        if not stocks or item.warehouse not in stocks:
            raise HTTPException(
                status_code=404,
                detail='Товар не найден на складе'
            )
        
        if stocks[item.warehouse] < item.quantity:
            raise HTTPException(
                status_code=400,
                detail='Недостаточное количество товара на складе'
            )
        
        # Выполняем списание
        try:
            manager.remove_from_warehouse(
                item.sku,
                warehouse.value,
                item.quantity
            )
            
            # Создаем запись операции
            operations_service = get_operations_service()
            operations_service.create_single_operation(
                operation_type=OperationType.STOCK_OUT_MANUAL,
                warehouse_id=warehouse.value,
                sku=item.sku,
                quantity=item.quantity,
                user_email=current_user.email,
            )
            
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        
        return JSONResponse({
            'status': 'success',
            'message': f'Успешно списано {item.quantity} шт. со склада {warehouse.value}'
        })
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post('/add/', summary='Пополнение товара на складе')
async def add_to_stock(
    item: AddItem,
    manager: manager.InventoryManager = Depends(manager.get_manager),
    current_user: User = Depends(deps.get_current_user_optional)
):
    '''Пополняет указанное количество товара на определенном складе.'''
    try:
        # Проверяем корректность склада
        try:
            warehouse = Warehouses(item.warehouse)
        except ValueError:
            raise HTTPException(status_code=400, detail='Неверное значение склада')
        
        # Проверяем существование товара
        stocks = manager.get_stock_by_sku(item.sku)
        
        # Выполняем пополнение
        manager.add_to_warehouse(
            item.sku,
            warehouse.value,
            item.quantity
        )
        
        # Создаем запись операции
        operations_service = get_operations_service()
        operations_service.create_single_operation(
            operation_type=OperationType.STOCK_IN,
            warehouse_id=warehouse.value,
            sku=item.sku,
            quantity=item.quantity,
            user_email=current_user.email,
        )
        
        return JSONResponse({
            'status': 'success',
            'message': f'Успешно добавлено {item.quantity} шт. на склад {warehouse.value}'
        })
        
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get(
    "/export/stock-with-sales/",
    summary="Экспорт остатков со статистикой продаж",
    response_class=StreamingResponse,
)
async def export_stock_with_sales(
    request: Request,
    skus: Optional[List[str]] = Query(None, description="Список SKU для экспорта"),
    manager: manager.InventoryManager = Depends(manager.get_manager),
):
    mimetypes.init()
    strict_map = mimetypes.types_map.copy()
    common_map = mimetypes.common_types.copy()
    mimetypes.types_map = {
        True: strict_map,
        False: common_map,
    }
    mimetypes.types_map[True][".webp"] = "image/webp"

    """Возвращает XLSX-файл с остатками, картинками 150×150px и статистикой продаж за 15/30/60 дней."""
    # получаем DF без image и параллельный список bytes
    df, images = manager.get_stock_report()
    
    # получаем статистику продаж
    sales_stats = manager.get_sales_statistics(skus)

    # фильтрация, если skus заданы
    if skus:
        mask = df["sku"].isin(skus)
        df = df[mask].reset_index(drop=True)
        images = [img for keep, img in zip(mask, images) if keep]

    # добавляем колонки со статистикой продаж
    df['Продажи за 15 дней'] = df['sku'].map(lambda x: sales_stats.get(x, {}).get('15d', 0))
    df['Продажи за 30 дней'] = df['sku'].map(lambda x: sales_stats.get(x, {}).get('30d', 0))
    df['Продажи за 60 дней'] = df['sku'].map(lambda x: sales_stats.get(x, {}).get('60d', 0))

    # добавляем колонку с абсолютным URL изображения
    # Проверяем заголовки прокси для определения правильной схемы
    scheme = request.headers.get('x-forwarded-proto', 'https' if request.url.scheme == 'https' else 'http')
    host = request.headers.get('host', str(request.base_url.hostname))
    base_url = f"{scheme}://{host}"
    df['URL изображения'] = df['sku'].map(lambda x: f"{base_url}/api/products/{x}/image/original")

    # переименования для читабельности
    column_renames = {
        "sku": "SKU",
        "eans": "EAN коды",
        "name": "Наименование",
        "total": "Общий остаток"
    }
    # добавляем названия складов
    for wh in Warehouses:
        column_renames[f"warehouse_{wh.value}"] = f"Склад {wh.value}"
    
    df = df.rename(columns=column_renames)

    # вставляем пустую колонку-«держатель» под изображения
    df.insert(1, "Изображение", "")

    # порядок колонок - сначала общий остаток, потом по складам
    cols = ["SKU", "Изображение", "URL изображения", "Наименование", "EAN коды", "Общий остаток"]
    
    # Сначала добавляем склад B
    cols.append(f"Склад {Warehouses.B.value}")
    # Затем добавляем все остальные склады
    for wh in Warehouses:
        if wh != Warehouses.B:
            cols.append(f"Склад {wh.value}")
            
    cols.extend(["Продажи за 15 дней", "Продажи за 30 дней", "Продажи за 60 дней"])
    df = df[cols]

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # пишем данные (картинки пока пустые)
        df.to_excel(writer, index=False, sheet_name="Остатки")
        ws = writer.sheets["Остатки"]

        # находим букву колонки «Изображение»
        img_col_idx    = df.columns.get_loc("Изображение") + 1
        img_col_letter = get_column_letter(img_col_idx)

        # 1) форматируем заголовки и ширину
        for idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.font = cell.font.copy(bold=True)
            if col == "Изображение":
                ws.column_dimensions[img_col_letter].width = 22  # ~150px
            elif col == "URL изображения":
                ws.column_dimensions[get_column_letter(idx)].width = 60  # для длинных URL
            else:
                max_len = max(df[col].astype(str).map(len).max(), len(col))
                ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)

        # 2) вставка картинок 150×150 и высота строк
        for row_idx, img_bytes in enumerate(images, start=2):
            if not img_bytes:
                continue
            try:
                img = XLImage(BytesIO(img_bytes))
                img.width  = 150
                img.height = 150
                ws.row_dimensions[row_idx].height = 115  # под ~150px

                ws.add_image(img, f"{img_col_letter}{row_idx}")
            except Exception as e:
                logging.error(f"Не удалось вставить изображение в строке {row_idx}: {e}")

    buf.seek(0)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"stock_report_with_sales_{ts}.xlsx"
    headers = {
        "Content-Disposition": (
            f'attachment; filename="{filename}"; '
            f"filename*=UTF-8''{filename}"
        )
    }
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers
    )

@router.get('/export/stock-with-sales/no-images/', summary='Экспорт остатков со статистикой продаж без изображений')
async def export_stock_with_sales_no_images(
    request: Request,
    skus: Optional[List[str]] = Query(None, description="Список SKU для экспорта"),
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    '''Возвращает XLSX-файл с остатками по складам, общим остатком и статистикой продаж за 15/30/60 дней, без изображений.'''
    df, _ = manager.get_stock_report()  # Игнорируем список изображений
    
    # получаем статистику продаж
    sales_stats = manager.get_sales_statistics(skus)

    # фильтрация, если skus заданы
    if skus:
        mask = df["sku"].isin(skus)
        df = df[mask].reset_index(drop=True)

    # добавляем колонки со статистикой продаж
    df['Продажи за 15 дней'] = df['sku'].map(lambda x: sales_stats.get(x, {}).get('15d', 0))
    df['Продажи за 30 дней'] = df['sku'].map(lambda x: sales_stats.get(x, {}).get('30d', 0))
    df['Продажи за 60 дней'] = df['sku'].map(lambda x: sales_stats.get(x, {}).get('60d', 0))

    # добавляем колонку с абсолютным URL изображения
    # Проверяем заголовки прокси для определения правильной схемы
    scheme = request.headers.get('x-forwarded-proto', 'https' if request.url.scheme == 'https' else 'http')
    host = request.headers.get('host', str(request.base_url.hostname))
    base_url = f"{scheme}://{host}"
    df['URL изображения'] = df['sku'].map(lambda x: f"{base_url}/api/products/{x}/image/original")
    
    # Русские названия столбцов
    column_renames = {
        "sku": "SKU",
        "eans": "EAN коды",
        "name": "Наименование",
        "total": "Общий остаток"
    }
    # добавляем названия складов
    for wh in Warehouses:
        column_renames[f"warehouse_{wh.value}"] = f"Склад {wh.value}"
    
    df = df.rename(columns=column_renames)
    
    # порядок колонок - сначала общий остаток, потом по складам
    cols = ["SKU", "URL изображения", "Наименование", "EAN коды", "Общий остаток"]
    
    # Сначала добавляем склад B
    cols.append(f"Склад {Warehouses.B.value}")
    # Затем добавляем все остальные склады
    for wh in Warehouses:
        if wh != Warehouses.B:
            cols.append(f"Склад {wh.value}")
            
    cols.extend(["Продажи за 15 дней", "Продажи за 30 дней", "Продажи за 60 дней"])
    df = df[cols]
    
    # Создаем Excel-файл
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Остатки')
        worksheet = writer.sheets['Остатки']
        
        # Форматируем заголовки и ширины
        for idx, col in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=idx)
            cell.font = cell.font.copy(bold=True)
            
            # Устанавливаем ширину колонок
            if col == "URL изображения":
                worksheet.column_dimensions[get_column_letter(idx)].width = 60  # для длинных URL
            else:
                max_len = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 50)
    
    buffer.seek(0)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"stock_report_with_sales_no_images_{ts}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
    }
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )

@router.post('/sale-from-order/', summary='Списание товаров из заказа')
async def sale_from_order(
    sale_data: SaleFromOrder,
    manager: manager.InventoryManager = Depends(manager.get_manager),
    db: AsyncSession = Depends(deps.get_async_session),
    current_user: User = Depends(deps.get_current_user_optional)
):
    '''Списывает товары из заказа как продажу только если все товары есть в наличии.'''
    try:
        # Сначала проверяем наличие всех товаров
        for item in sale_data.line_items:
            if 'external_id' in item and item['external_id']:
                stocks = manager.get_stock_by_sku(item['external_id'])
                if not stocks or sale_data.warehouse not in stocks or stocks[sale_data.warehouse] < 1:
                    return JSONResponse(
                        status_code=400,
                        content={
                            'status': 'error',
                            'message': f'Товар {item["external_id"]} отсутствует на складе {sale_data.warehouse}'
                        }
                    )

        products = []
        # Если все товары есть в наличии, выполняем списание
        operations_service = get_operations_service()
        for item in sale_data.line_items:
            if 'external_id' in item and item['external_id']:
                try:
                    manager.remove_as_sale(
                        sku=item['external_id'],
                        warehouse=sale_data.warehouse,
                        quantity=1  # Списываем по одной единице для каждой позиции
                    )
                    products.append({
                        'sku': item['external_id'],
                        'quantity': 1
                    })
                    # Создаем запись операции для каждого списанного товара
                    
                except Exception as e:
                    return JSONResponse(
                        status_code=400,
                        content={
                            'status': 'error',
                            'message': f'Ошибка при списании товара {item["external_id"]}: {str(e)}'
                        }
                    )

        operations_service.create_order_operation(
            warehouse_id=sale_data.warehouse,
            user_email=current_user.email,
            order_id=sale_data.order_id,
            comment=f"Списание товара выполнено через кнопку 'списать'\n",
            products_data=products
        )
        update_stmt = update(AllegroOrder).where(
            AllegroOrder.id == sale_data.order_id
        ).values(is_stock_updated=True)
        
        await db.exec(update_stmt)
        await db.commit()
        
        return JSONResponse({
            'status': 'success',
            'message': f'Успешно списаны товары из заказа {sale_data.order_id}'
        })
        
    except Exception as e:
        await db.rollback()
        return JSONResponse(
            status_code=500,
            content={
                'status': 'error',
                'message': f'Ошибка при обработке заказа: {str(e)}'
            }
        )

@router.get('/compress-images/', summary='Сжатие изображений')
async def compress_images(
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    result = manager.compress_all_product_images()
    return JSONResponse({'status': 'success', 'result': result})

@router.get('/decompress-images/', summary='Декомпрессия изображений')
async def decompress_images(
    manager: manager.InventoryManager = Depends(manager.get_manager)
):
    """
    Декомпрессирует все изображения продуктов до размера 600x600 в формат JPEG.
    Записывает декомпрессированные изображения в поле original_image и создает URL.
    """
    result = manager.decompress_all_product_images()
    return JSONResponse({'status': 'success', 'result': result})

@router.post('/sync-wix/', summary='Синхронизация количества товаров с Wix')
async def sync_wix_inventory(
    background_tasks: BackgroundTasks,
):
    """
    Запускает синхронизацию количества товаров между локальной базой данных и Wix.
    
    Процесс:
    1. Получает все SKU и их количество из локальной базы данных
    2. Находит соответствующие товары в Wix по SKU
    3. Обновляет количество товаров в Wix до соответствия локальной базе
    
    Returns:
        JSONResponse: Результат запуска синхронизации
    """

    try:
        from app.celery_app import launch_wix_sync
        
        # Запускаем синхронизацию в фоновом режиме
        result = launch_wix_sync()
        
        return JSONResponse({
            "status": "success",
            "message": "Синхронизация с Wix запущена",
            "task_id": result.id,
            "task_status": "PENDING"
        })
        
    except Exception as e:
        logging.error(f"Ошибка при запуске синхронизации Wix: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Ошибка при запуске синхронизации: {str(e)}"
        )


@router.get('/sync-wix/status/{task_id}', summary='Статус синхронизации Wix')
async def get_wix_sync_status(
    task_id: str,
    current_user: User = Depends(deps.get_current_user_optional)
):
    """
    Получает статус выполнения задачи синхронизации с Wix.
    
    Args:
        task_id: ID задачи Celery
        
    Returns:
        JSONResponse: Статус задачи и результат выполнения
    """
    if not current_user:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    
    try:
        from celery.result import AsyncResult
        
        # Получаем результат задачи
        result = AsyncResult(task_id)
        
        response_data = {
            "task_id": task_id,
            "status": result.status,
            "ready": result.ready()
        }
        
        # Если задача завершена, добавляем результат
        if result.ready():
            if result.successful():
                response_data["result"] = result.result
            else:
                response_data["error"] = str(result.info)
        
        return JSONResponse(response_data)
        
    except Exception as e:
        logging.error(f"Ошибка при получении статуса задачи {task_id}: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Ошибка при получении статуса: {str(e)}"
        )
